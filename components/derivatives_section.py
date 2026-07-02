from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
import hashlib
import re
import sys

import pandas as pd
import streamlit as st
import plotly.graph_objects as go

DERIV_ROOT = Path(__file__).resolve().parents[1] / "derivatives_app"
_DERIV_IMPORT_ERROR = None

try:
    from derivatives_src.chart_registry import CHARTS, blocks, charts_for_block
    from derivatives_src.io.bde_table import test_bde_url
    from derivatives_src.io.bcch_api import BCCHClient as DerivBCCHClient
    from derivatives_src.series_registry import SeriesRegistry
except Exception as exc:  # defensivo: no debe botar toda la app si falta el módulo derivado
    _DERIV_IMPORT_ERROR = exc
    CHARTS = []

    def blocks() -> list[str]:
        return []

    def charts_for_block(block: str) -> list:
        return []

    def test_bde_url(url: str) -> tuple[int, int]:
        raise RuntimeError(str(_DERIV_IMPORT_ERROR))

    DerivBCCHClient = None
    SeriesRegistry = None

APP_VERSION = "v38_deriv_menu_summary_report"
DEFAULT_MAP = DERIV_ROOT / "config" / "series_map.csv"
TEST_BDE_URL = "https://si3.bcentral.cl/Siete/ES/Siete/Cuadro/CAP_DERYSPOT/MN_DERYSPOT/DER_RES_POS_01"

_DERIV_USER = ""
_DERIV_PASS = ""


def set_derivatives_credentials(user: str, password: str) -> None:
    global _DERIV_USER, _DERIV_PASS
    _DERIV_USER = user or ""
    _DERIV_PASS = password or ""

@st.cache_data(show_spinner=False)
def load_registry(path: str):
    if SeriesRegistry is None:
        raise RuntimeError(f"Módulo derivados no disponible: {_DERIV_IMPORT_ERROR}")
    return SeriesRegistry.from_csv(path)


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def cached_build_dataframe(module_name: str, start_date: str, end_date: str, map_path: str, map_version: str, app_version: str, user: str, password_hash: str, _password: str) -> pd.DataFrame:
    if SeriesRegistry is None:
        return pd.DataFrame()
    registry = SeriesRegistry.from_csv(map_path)
    spec = next((c for c in CHARTS if c.module_name == module_name), None)
    if spec is None:
        return pd.DataFrame()
    client = DerivBCCHClient(user=user, password=_password) if (DerivBCCHClient is not None and user and _password) else None
    try:
        return spec.build_dataframe(client, registry, start_date, end_date, False)
    except Exception as exc:
        # No botar la app completa por un gráfico específico de derivados.
        return pd.DataFrame({"error": [str(exc)], "chart_module": [module_name]})


@st.cache_data(show_spinner=True, ttl=60 * 60 * 6)
def bde_test_count() -> tuple[int, int]:
    return test_bde_url(TEST_BDE_URL)


def soft_divider() -> None:
    st.markdown('<hr class="soft-divider" />', unsafe_allow_html=True)


def section_title(text: str) -> None:
    st.markdown(f'<div class="section-title">{text}</div>', unsafe_allow_html=True)


def chart_title(text: str, unit: str = "Millones de USD") -> None:
    st.markdown(
        f"""
        <div class="chart-title-card">
            <div class="chart-title">{text}</div>
            <div class="chart-unit">Unidad: {unit}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def get_d01_context(df: pd.DataFrame) -> dict:
    """Controles compartidos para D01.3 y D01.4."""
    from derivatives_src.charts.d01_waterfall_variacion import available_metrics, available_months, month_label, METRIC_STOCK

    metrics = available_metrics(df)
    default_metric = METRIC_STOCK if METRIC_STOCK in metrics else (metrics[0] if metrics else METRIC_STOCK)

    st.markdown("#### Parámetros de comparación")
    c0, c1, c2 = st.columns(3)
    with c0:
        metric = st.selectbox(
            "Tipo de monto",
            options=metrics,
            index=metrics.index(default_metric) if default_metric in metrics else 0,
            key="d01_compare_metric",
        )

    months = available_months(df, metric)
    if len(months) < 2:
        return {"metric": metric, "base_month": None, "compare_month": None}

    with c1:
        base_month = st.selectbox(
            "Mes base",
            options=months,
            index=max(len(months) - 2, 0),
            format_func=month_label,
            key="d01_var_base_month",
        )
    with c2:
        compare_month = st.selectbox(
            "Mes comparación",
            options=months,
            index=len(months) - 1,
            format_func=month_label,
            key="d01_var_compare_month",
        )

    ctx = {"metric": metric, "base_month": base_month, "compare_month": compare_month}
    st.session_state["d01_compare_context"] = ctx
    return ctx




def get_month_compare_context(df: pd.DataFrame, prefix: str, title: str = "Parámetros de comparación") -> dict:
    from derivatives_src.charts.d01_waterfall_variacion import month_label

    months = sorted(pd.to_datetime(df["date"]).dropna().unique()) if (not df.empty and "date" in df.columns) else []
    st.markdown(f"#### {title}")
    if len(months) < 2:
        return {"base_month": None, "compare_month": None}
    c1, c2 = st.columns(2)
    with c1:
        base_month = st.selectbox(
            "Mes base", options=months, index=max(len(months) - 2, 0), format_func=month_label, key=f"{prefix}_base_month"
        )
    with c2:
        compare_month = st.selectbox(
            "Mes comparación", options=months, index=len(months) - 1, format_func=month_label, key=f"{prefix}_compare_month"
        )
    return {"base_month": base_month, "compare_month": compare_month}

def build_df_for_spec(spec, start: date, end: date, active_map_path: str) -> pd.DataFrame:
    preloaded = st.session_state.get("deriv_preloaded_frames", {})
    if isinstance(preloaded, dict) and spec.module_name in preloaded:
        df = preloaded.get(spec.module_name)
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()

    map_path = Path(active_map_path)
    map_version = str(map_path.stat().st_mtime_ns) if map_path.exists() else "0"
    return cached_build_dataframe(
        spec.module_name,
        str(start),
        str(end),
        active_map_path,
        map_version,
        APP_VERSION,
        _DERIV_USER,
        _password_hash(_DERIV_PASS),
        _DERIV_PASS,
    )



def download_csv(df: pd.DataFrame, key: str, filename: str) -> None:
    if not df.empty:
        st.download_button(
            "Descargar CSV",
            data=df.to_csv(index=False, encoding="utf-8-sig"),
            file_name=filename,
            mime="text/csv",
            key=key,
        )


def _safe_excel_sheet_name(name: str, used: set[str]) -> str:
    """Excel limita nombres a 31 caracteres y prohíbe algunos símbolos."""
    clean = re.sub(r"[\\/\?\*\[\]:]", "_", str(name)).strip()
    clean = re.sub(r"\s+", "_", clean) or "Hoja"
    base = clean[:31]
    candidate = base
    counter = 1
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def _prepare_excel_frame(df: pd.DataFrame, spec) -> pd.DataFrame:
    """Agrega metadata útil y evita problemas típicos al exportar a Excel."""
    if df is None or df.empty:
        return pd.DataFrame(
            {
                "block": [spec.block],
                "chart_title": [spec.title],
                "chart_module": [spec.module_name],
                "status": ["sin datos"],
            }
        )

    out = df.copy()
    out.insert(0, "chart_module", spec.module_name)
    out.insert(1, "chart_title", spec.title)
    out.insert(2, "block", spec.block)

    # Excel interpreta texto que parte con '=' como fórmula. Para exportación segura,
    # protegemos columnas de texto sin alterar números ni fechas.
    for col in out.select_dtypes(include=["object"]).columns:
        out[col] = out[col].map(lambda x: "'" + x if isinstance(x, str) and x.startswith("=") else x)
    return out


def _format_export_workbook(writer, sheet_names: list[str]) -> None:
    """Formato liviano para que el workbook sea legible sin ralentizar la app."""
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = writer.book
    header_fill = PatternFill(fill_type="solid", fgColor="5B2A86")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_cells in ws.columns:
            header = str(col_cells[0].value or "")
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[:200])
            width = min(max(max_len + 2, min(max(len(header) + 2, 18), 18)), 38)
            ws.column_dimensions[col_cells[0].column_letter].width = width
        # Formato fecha para columnas llamadas date/fecha.
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).lower() in {"date", "fecha"}:
                for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    row[0].number_format = "yyyy-mm-dd"


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def build_powerbi_excel_bytes(start_date: str, end_date: str, map_path: str, map_version: str, app_version: str, user: str, _password: str) -> bytes:
    """Construye un Excel con una hoja por gráfico, no solo por bloque visible."""
    registry = SeriesRegistry.from_csv(map_path)
    built: list[tuple] = []
    index_rows: list[dict] = []
    used_sheets: set[str] = set()

    for idx, spec in enumerate(CHARTS, start=1):
        sheet_name = _safe_excel_sheet_name(f"{idx:02d}_{spec.module_name}", used_sheets)
        try:
            client = DerivBCCHClient(user=user, password=_password) if user and _password else None
            df = spec.build_dataframe(client, registry, start_date, end_date, False)
            status = "ok" if not df.empty else "sin datos"
            rows = int(len(df)) if df is not None else 0
            export_df = _prepare_excel_frame(df, spec)
            error = ""
        except Exception as exc:
            status = "error"
            rows = 0
            error = str(exc)
            export_df = pd.DataFrame(
                {
                    "block": [spec.block],
                    "chart_title": [spec.title],
                    "chart_module": [spec.module_name],
                    "status": ["error"],
                    "error": [error],
                }
            )

        built.append((sheet_name, export_df))
        index_rows.append(
            {
                "sheet": sheet_name,
                "block": spec.block,
                "chart_title": spec.title,
                "chart_module": spec.module_name,
                "chart_id": spec.chart_id or "",
                "rows": rows,
                "status": status,
                "error": error,
            }
        )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        index_df = pd.DataFrame(index_rows)
        index_df.to_excel(writer, sheet_name="INDICE", index=False)
        written_sheets = ["INDICE"]
        for sheet_name, export_df in built:
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            written_sheets.append(sheet_name)
        _format_export_workbook(writer, written_sheets)

    return output.getvalue()


def build_powerbi_excel_for_download(start: date, end: date, active_map_path: str) -> bytes:
    """Construye Excel global de derivados con las credenciales activas de SieteWS."""
    map_path = Path(active_map_path)
    map_version = str(map_path.stat().st_mtime_ns) if map_path.exists() else "0"
    return build_powerbi_excel_bytes(str(start), str(end), active_map_path, map_version, APP_VERSION, _DERIV_USER, _DERIV_PASS)


def render_chart(spec, start: date, end: date, active_map_path: str, add_divider: bool = True) -> pd.DataFrame:
    chart_title(spec.title)
    try:
        df = build_df_for_spec(spec, start, end, active_map_path)
        fig = spec.build_figure(df)
        render_deriv_plotly(fig)
        download_csv(df, f"download_{spec.module_name}", f"{spec.module_name}.csv")
        if add_divider:
            soft_divider()
        return df
    except Exception as exc:
        st.error(f"No se pudo construir {spec.title}: {exc}")
        if add_divider:
            soft_divider()
        return pd.DataFrame()


def render_d01_variation_pair(spec_var, spec_comp, start: date, end: date, active_map_path: str) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    try:
        df_var = build_df_for_spec(spec_var, start, end, active_map_path)
        df_comp = build_df_for_spec(spec_comp, start, end, active_map_path)

        # Controles compartidos arriba para que ambos gráficos queden alineados.
        if not df_var.empty:
            ctx = get_d01_context(df_var)
        else:
            ctx = {"metric": None, "base_month": None, "compare_month": None}

        from derivatives_src.charts.d01_variacion_componentes import available_markets
        markets = available_markets(df_comp)
        default_market = "Tipos de cambio" if "Tipos de cambio" in markets else (markets[0] if markets else "Tipos de cambio")
        market = st.selectbox(
            "Mercado a descomponer",
            options=markets,
            index=markets.index(default_market) if default_market in markets else 0,
            key="d01_component_market",
        )

        c_left, c_right = st.columns(2, gap="large")
        with c_left:
            chart_title(spec_var.title)
            fig_var = spec_var.build_figure(df_var, ctx.get("base_month"), ctx.get("compare_month"), ctx.get("metric"))
            render_deriv_plotly(fig_var)
            download_csv(df_var, f"download_{spec_var.module_name}", f"{spec_var.module_name}.csv")
        with c_right:
            chart_title(spec_comp.title)
            fig_comp = spec_comp.build_figure(df_comp, ctx.get("base_month"), ctx.get("compare_month"), ctx.get("metric"), market)
            render_deriv_plotly(fig_comp)
            download_csv(df_comp, f"download_{spec_comp.module_name}", f"{spec_comp.module_name}.csv")
        soft_divider()
        frames.extend([df_var, df_comp])
    except Exception as exc:
        st.error(f"No se pudo construir comparación D01: {exc}")
        soft_divider()
    return frames



def render_d02_variacion_apertura(spec_var, start: date, end: date, active_map_path: str) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    try:
        df_var = build_df_for_spec(spec_var, start, end, active_map_path)
        from derivatives_src.charts import d02_fx_variacion_apertura_sector as varmod

        st.markdown("#### Parámetros de comparación")
        openings = varmod.available_openings(df_var)
        default_opening = "Comprador" if "Comprador" in openings else (openings[0] if openings else "Comprador")
        c0, c1, c2 = st.columns(3)
        with c0:
            apertura = st.selectbox(
                "Apertura",
                options=openings,
                index=openings.index(default_opening) if default_opening in openings else 0,
                key="d02_var_apertura",
            )
        months = varmod.available_months(df_var, apertura)
        if len(months) < 2:
            ctx = {"base_month": None, "compare_month": None}
            st.warning("No hay meses suficientes para comparar la apertura seleccionada.")
        else:
            from derivatives_src.charts.d01_waterfall_variacion import month_label
            with c1:
                base_month = st.selectbox("Mes base", options=months, index=max(len(months) - 2, 0), format_func=month_label, key="d02_var_base_month")
            with c2:
                compare_month = st.selectbox("Mes comparación", options=months, index=len(months) - 1, format_func=month_label, key="d02_var_compare_month")
            ctx = {"base_month": base_month, "compare_month": compare_month}

        c_left, c_right = st.columns(2, gap="large")
        with c_left:
            chart_title("Variación por sector")
            fig_sector = varmod.build_figure_sector(df_var, ctx.get("base_month"), ctx.get("compare_month"), apertura)
            render_deriv_plotly(fig_sector)
        with c_right:
            chart_title("Ranking de contribución")
            fig_rank = varmod.build_figure_ranking(df_var, ctx.get("base_month"), ctx.get("compare_month"), apertura)
            render_deriv_plotly(fig_rank)

        download_csv(df_var, f"download_{spec_var.module_name}", f"{spec_var.module_name}.csv")
        soft_divider()
        frames.append(df_var)
    except Exception as exc:
        st.error(f"No se pudo construir variación D02 por apertura y sector: {exc}")
        soft_divider()
    return frames

def render_d02_buyer_seller_pair(spec_buy, spec_sell, start: date, end: date, active_map_path: str) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    try:
        df_buy = build_df_for_spec(spec_buy, start, end, active_map_path)
        df_sell = build_df_for_spec(spec_sell, start, end, active_map_path)
        df_ctx = df_buy if not df_buy.empty else df_sell
        ctx = get_month_compare_context(df_ctx, "d02_buy_sell", "Parámetros de comparación")

        c_left, c_right = st.columns(2, gap="large")
        with c_left:
            chart_title(spec_buy.title)
            fig_buy = spec_buy.build_figure(df_buy, ctx.get("base_month"), ctx.get("compare_month"))
            render_deriv_plotly(fig_buy)
            download_csv(df_buy, f"download_{spec_buy.module_name}", f"{spec_buy.module_name}.csv")
        with c_right:
            chart_title(spec_sell.title)
            fig_sell = spec_sell.build_figure(df_sell, ctx.get("base_month"), ctx.get("compare_month"))
            render_deriv_plotly(fig_sell)
            download_csv(df_sell, f"download_{spec_sell.module_name}", f"{spec_sell.module_name}.csv")
        soft_divider()
        frames.extend([df_buy, df_sell])
    except Exception as exc:
        st.error(f"No se pudo construir comparación D02 comprador/vendedor: {exc}")
        soft_divider()
    return frames




def render_d02_spot_compras_ventas_variacion(spec_var, start: date, end: date, active_map_path: str) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    try:
        df_var = build_df_for_spec(spec_var, start, end, active_map_path)
        from derivatives_src.charts import d02_spot_compras_ventas_variacion as varmod
        from derivatives_src.charts.d01_waterfall_variacion import month_label

        st.markdown("#### Parámetros de comparación spot compras/ventas USD")
        metrics = varmod.available_metrics(df_var)
        default_metric = "Compras USD" if "Compras USD" in metrics else (metrics[0] if metrics else "Compras USD")
        c0, c1, c2 = st.columns(3)

        with c0:
            metrica = st.selectbox(
                "Métrica",
                options=metrics,
                index=metrics.index(default_metric) if default_metric in metrics else 0,
                key="d02_spot_cv_metrica",
            )

        months = varmod.available_months(df_var, metrica)
        if len(months) < 2:
            ctx = {"base_month": None, "compare_month": None}
            st.warning("No hay meses suficientes para comparar la métrica seleccionada.")
        else:
            with c1:
                base_month = st.selectbox(
                    "Mes base",
                    options=months,
                    index=max(len(months) - 2, 0),
                    format_func=month_label,
                    key="d02_spot_cv_base_month",
                )
            with c2:
                compare_month = st.selectbox(
                    "Mes comparación",
                    options=months,
                    index=len(months) - 1,
                    format_func=month_label,
                    key="d02_spot_cv_compare_month",
                )
            ctx = {"base_month": base_month, "compare_month": compare_month}

        c_left, c_right = st.columns(2, gap="large")
        with c_left:
            chart_title("Variación por sector")
            fig_sector = varmod.build_figure_sector(df_var, ctx.get("base_month"), ctx.get("compare_month"), metrica)
            render_deriv_plotly(fig_sector)
        with c_right:
            chart_title("Ranking de contribución")
            fig_rank = varmod.build_figure_ranking(df_var, ctx.get("base_month"), ctx.get("compare_month"), metrica)
            render_deriv_plotly(fig_rank)

        download_csv(df_var, f"download_{spec_var.module_name}", f"{spec_var.module_name}.csv")
        soft_divider()
        frames.append(df_var)
    except Exception as exc:
        st.error(f"No se pudo construir comparación spot compras/ventas USD: {exc}")
        soft_divider()
    return frames




def render_d05_ufclp_transados(specs: list, start: date, end: date, active_map_path: str) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    spec_map = {s.module_name: s for s in specs}

    try:
        from derivatives_src.charts.d05_ufclp_common import UNIT_USD, available_aperturas
        from derivatives_src.charts import d05_ufclp_forward_sector as fwd_mod, d05_ufclp_swap_sector as swap_mod

        options = available_aperturas(pd.DataFrame())
        st.markdown("#### Parámetros — UF/CLP por instrumento")
        apertura = st.selectbox(
            "Compra/venta UF",
            options=options,
            index=options.index("Total") if "Total" in options else 0,
            key="d05_ufclp_apertura",
        )

        c1, c2 = st.columns(2, gap="large")

        if "d05_ufclp_forward_sector" in spec_map:
            spec = spec_map["d05_ufclp_forward_sector"]
            df_fwd = build_df_for_spec(spec, start, end, active_map_path)
            with c1:
                chart_title(spec.title, unit=UNIT_USD)
                fig = fwd_mod.build_figure(df_fwd, apertura)
                render_deriv_plotly(fig)
                download_csv(df_fwd, f"download_{spec.module_name}", f"{spec.module_name}.csv")
            frames.append(df_fwd)

        if "d05_ufclp_swap_sector" in spec_map:
            spec = spec_map["d05_ufclp_swap_sector"]
            df_swap = build_df_for_spec(spec, start, end, active_map_path)
            with c2:
                chart_title(spec.title, unit=UNIT_USD)
                fig = swap_mod.build_figure(df_swap, apertura)
                render_deriv_plotly(fig)
                download_csv(df_swap, f"download_{spec.module_name}", f"{spec.module_name}.csv")
            frames.append(df_swap)

        soft_divider()
    except Exception as exc:
        st.error(f"No se pudo construir bloque UF/CLP transados por instrumento: {exc}")
        soft_divider()

    return frames


def render_d04_spc_nominal(specs: list, start: date, end: date, active_map_path: str) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    spec_map = {s.module_name: s for s in specs}

    try:
        from derivatives_src.charts.d04_spc_nominal_common import UNIT_CLP
        from derivatives_src.charts.d01_waterfall_variacion import month_label
        from derivatives_src.charts import (
            d04_spc_nominal_sector as sector_mod,
            d04_spc_nominal_plazo as plazo_mod,
            d04_spc_nominal_variacion_plazo as var_mod,
        )

        # 1) Sector de contraparte
        if "d04_spc_nominal_sector" in spec_map:
            spec = spec_map["d04_spc_nominal_sector"]
            df_sector = build_df_for_spec(spec, start, end, active_map_path)

            st.markdown("#### Parámetros — sector de contraparte")
            aperturas = sector_mod.available_aperturas(df_sector)
            plazos = sector_mod.available_plazos(df_sector)
            c0, c1 = st.columns(2)
            with c0:
                apertura_sector = st.selectbox(
                    "Compra/venta",
                    options=aperturas,
                    index=aperturas.index("Total") if "Total" in aperturas else 0,
                    key="d04_spc_sector_apertura",
                )
            with c1:
                plazo_sector = st.selectbox(
                    "Plazo contractual",
                    options=plazos,
                    index=0,
                    key="d04_spc_sector_plazo",
                )

            chart_title(spec.title, unit=UNIT_CLP)
            fig_sector = sector_mod.build_figure(df_sector, apertura_sector, plazo_sector)
            render_deriv_plotly(fig_sector)
            download_csv(df_sector, f"download_{spec.module_name}", f"{spec.module_name}.csv")
            soft_divider()
            frames.append(df_sector)

        # 2) Plazo contractual
        if "d04_spc_nominal_plazo" in spec_map:
            spec = spec_map["d04_spc_nominal_plazo"]
            df_plazo = build_df_for_spec(spec, start, end, active_map_path)

            st.markdown("#### Parámetros — plazo contractual")
            aperturas = plazo_mod.available_aperturas(df_plazo)
            apertura_plazo = st.selectbox(
                "Compra/venta",
                options=aperturas,
                index=aperturas.index("Total") if "Total" in aperturas else 0,
                key="d04_spc_plazo_apertura",
            )

            chart_title(spec.title, unit=UNIT_CLP)
            fig_plazo = plazo_mod.build_figure(df_plazo, apertura_plazo)
            render_deriv_plotly(fig_plazo)
            download_csv(df_plazo, f"download_{spec.module_name}", f"{spec.module_name}.csv")
            soft_divider()
            frames.append(df_plazo)

        # 3) Variación por plazo contractual
        if "d04_spc_nominal_variacion_plazo" in spec_map:
            spec = spec_map["d04_spc_nominal_variacion_plazo"]
            df_var = build_df_for_spec(spec, start, end, active_map_path)

            st.markdown("#### Parámetros de comparación — SPC nominal")
            aperturas = var_mod.available_aperturas(df_var)
            default_apertura = "Total" if "Total" in aperturas else (aperturas[0] if aperturas else "Total")
            c0, c1, c2 = st.columns(3)
            with c0:
                apertura_var = st.selectbox(
                    "Compra/venta",
                    options=aperturas,
                    index=aperturas.index(default_apertura) if default_apertura in aperturas else 0,
                    key="d04_spc_var_apertura",
                )

            months = var_mod.available_months(df_var, apertura_var)
            if len(months) < 2:
                ctx = {"base_month": None, "compare_month": None}
                st.warning("No hay meses suficientes para comparar SPC nominal.")
            else:
                with c1:
                    base_month = st.selectbox(
                        "Mes base",
                        options=months,
                        index=max(len(months) - 2, 0),
                        format_func=month_label,
                        key="d04_spc_var_base_month",
                    )
                with c2:
                    compare_month = st.selectbox(
                        "Mes comparación",
                        options=months,
                        index=len(months) - 1,
                        format_func=month_label,
                        key="d04_spc_var_compare_month",
                    )
                ctx = {"base_month": base_month, "compare_month": compare_month}

            c_left, c_right = st.columns(2, gap="large")
            with c_left:
                chart_title("Variación SPC nominal por plazo contractual", unit=UNIT_CLP)
                fig_var = var_mod.build_figure_plazo(df_var, ctx.get("base_month"), ctx.get("compare_month"), apertura_var)
                render_deriv_plotly(fig_var)
            with c_right:
                chart_title("Ranking de contribución SPC nominal", unit=UNIT_CLP)
                fig_rank = var_mod.build_figure_ranking(df_var, ctx.get("base_month"), ctx.get("compare_month"), apertura_var)
                render_deriv_plotly(fig_rank)

            download_csv(df_var, f"download_{spec.module_name}", f"{spec.module_name}.csv")
            soft_divider()
            frames.append(df_var)

    except Exception as exc:
        st.error(f"No se pudo construir bloque SPC nominal: {exc}")
        soft_divider()

    return frames



def render_d04_spc_nominal_vigentes(specs: list, start: date, end: date, active_map_path: str) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    spec_map = {s.module_name: s for s in specs}

    try:
        from derivatives_src.charts.d04_spc_nominal_common import UNIT_CLP
        from derivatives_src.charts.d01_waterfall_variacion import month_label
        from derivatives_src.charts import (
            d04_spc_nominal_vigente_sector as sector_mod,
            d04_spc_nominal_vigente_plazo as plazo_mod,
            d04_spc_nominal_vigente_variacion_plazo as var_mod,
        )

        # 1) Sector de contraparte
        if "d04_spc_nominal_vigente_sector" in spec_map:
            spec = spec_map["d04_spc_nominal_vigente_sector"]
            df_sector = build_df_for_spec(spec, start, end, active_map_path)

            st.markdown("#### Parámetros — sector de contraparte")
            aperturas = sector_mod.available_aperturas(df_sector)
            plazos = sector_mod.available_plazos(df_sector)
            c0, c1 = st.columns(2)
            with c0:
                apertura_sector = st.selectbox(
                    "Compra/venta",
                    options=aperturas,
                    index=aperturas.index("Total") if "Total" in aperturas else 0,
                    key="d04_spc_vig_sector_apertura",
                )
            with c1:
                plazo_sector = st.selectbox(
                    "Plazo residual",
                    options=plazos,
                    index=0,
                    key="d04_spc_vig_sector_plazo",
                )

            chart_title(spec.title, unit=UNIT_CLP)
            fig_sector = sector_mod.build_figure(df_sector, apertura_sector, plazo_sector)
            render_deriv_plotly(fig_sector)
            download_csv(df_sector, f"download_{spec.module_name}", f"{spec.module_name}.csv")
            soft_divider()
            frames.append(df_sector)

        # 2) Plazo residual
        if "d04_spc_nominal_vigente_plazo" in spec_map:
            spec = spec_map["d04_spc_nominal_vigente_plazo"]
            df_plazo = build_df_for_spec(spec, start, end, active_map_path)

            st.markdown("#### Parámetros — plazo residual")
            aperturas = plazo_mod.available_aperturas(df_plazo)
            apertura_plazo = st.selectbox(
                "Compra/venta",
                options=aperturas,
                index=aperturas.index("Total") if "Total" in aperturas else 0,
                key="d04_spc_vig_plazo_apertura",
            )

            chart_title(spec.title, unit=UNIT_CLP)
            fig_plazo = plazo_mod.build_figure(df_plazo, apertura_plazo)
            render_deriv_plotly(fig_plazo)
            download_csv(df_plazo, f"download_{spec.module_name}", f"{spec.module_name}.csv")
            soft_divider()
            frames.append(df_plazo)

        # 3) Variación por plazo residual
        if "d04_spc_nominal_vigente_variacion_plazo" in spec_map:
            spec = spec_map["d04_spc_nominal_vigente_variacion_plazo"]
            df_var = build_df_for_spec(spec, start, end, active_map_path)

            st.markdown("#### Parámetros de comparación — SPC nominal vigente")
            aperturas = var_mod.available_aperturas(df_var)
            default_apertura = "Total" if "Total" in aperturas else (aperturas[0] if aperturas else "Total")
            c0, c1, c2 = st.columns(3)
            with c0:
                apertura_var = st.selectbox(
                    "Compra/venta",
                    options=aperturas,
                    index=aperturas.index(default_apertura) if default_apertura in aperturas else 0,
                    key="d04_spc_vig_var_apertura",
                )

            months = var_mod.available_months(df_var, apertura_var)
            if len(months) < 2:
                ctx = {"base_month": None, "compare_month": None}
                st.warning("No hay meses suficientes para comparar SPC nominal vigente.")
            else:
                with c1:
                    base_month = st.selectbox(
                        "Mes base",
                        options=months,
                        index=max(len(months) - 2, 0),
                        format_func=month_label,
                        key="d04_spc_vig_var_base_month",
                    )
                with c2:
                    compare_month = st.selectbox(
                        "Mes comparación",
                        options=months,
                        index=len(months) - 1,
                        format_func=month_label,
                        key="d04_spc_vig_var_compare_month",
                    )
                ctx = {"base_month": base_month, "compare_month": compare_month}

            c_left, c_right = st.columns(2, gap="large")
            with c_left:
                chart_title("Variación SPC nominal vigente por plazo residual", unit=UNIT_CLP)
                fig_var = var_mod.build_figure_plazo(df_var, ctx.get("base_month"), ctx.get("compare_month"), apertura_var)
                render_deriv_plotly(fig_var)
            with c_right:
                chart_title("Ranking de contribución SPC nominal vigente", unit=UNIT_CLP)
                fig_rank = var_mod.build_figure_ranking(df_var, ctx.get("base_month"), ctx.get("compare_month"), apertura_var)
                render_deriv_plotly(fig_rank)

            download_csv(df_var, f"download_{spec.module_name}", f"{spec.module_name}.csv")
            soft_divider()
            frames.append(df_var)

    except Exception as exc:
        st.error(f"No se pudo construir bloque SPC nominal vigente: {exc}")
        soft_divider()

    return frames


def render_d02_ndf_pair(spec_stock, spec_var, start: date, end: date, active_map_path: str, prefix: str, titulo: str) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    try:
        df_stock = build_df_for_spec(spec_stock, start, end, active_map_path)
        df_var = build_df_for_spec(spec_var, start, end, active_map_path)

        from derivatives_src.charts.d01_waterfall_variacion import month_label
        months = sorted(pd.to_datetime(df_var["date"]).dropna().unique()) if (not df_var.empty and "date" in df_var.columns) else []

        st.markdown(f"#### {titulo}")
        if len(months) < 2:
            ctx = {"base_month": None, "compare_month": None}
            st.warning("No hay meses suficientes para comparar esta serie NDF.")
        else:
            c1, c2 = st.columns(2)
            with c1:
                base_month = st.selectbox(
                    "Mes base",
                    options=months,
                    index=max(len(months) - 2, 0),
                    format_func=month_label,
                    key=f"{prefix}_base_month",
                )
            with c2:
                compare_month = st.selectbox(
                    "Mes comparación",
                    options=months,
                    index=len(months) - 1,
                    format_func=month_label,
                    key=f"{prefix}_compare_month",
                )
            ctx = {"base_month": base_month, "compare_month": compare_month}

        c_left, c_right = st.columns(2, gap="large")
        with c_left:
            chart_title(spec_stock.title)
            fig_stock = spec_stock.build_figure(df_stock)
            render_deriv_plotly(fig_stock)
            download_csv(df_stock, f"download_{spec_stock.module_name}", f"{spec_stock.module_name}.csv")
        with c_right:
            chart_title(spec_var.title)
            fig_var = spec_var.build_figure(df_var, ctx.get("base_month"), ctx.get("compare_month"))
            render_deriv_plotly(fig_var)
            download_csv(df_var, f"download_{spec_var.module_name}", f"{spec_var.module_name}.csv")

        soft_divider()
        frames.extend([df_stock, df_var])
    except Exception as exc:
        st.error(f"No se pudo construir bloque NDF {titulo}: {exc}")
        soft_divider()
    return frames





DERIV_DARK_COLORS = {
    "paper": "#0B1B4D",
    "plot": "#0F1E4A",
    "grid": "rgba(255,255,255,0.11)",
    "text": "#F5F7FA",
    "muted": "#B7C3D7",
    "primary": "#8B3DFF",
    "cyan": "#5AD7FF",
}


def apply_derivatives_dark_theme(fig):
    """Homologa gráficos del proyecto de derivados al tema oscuro CMF.

    Además fuerza la serie Total a una línea visible, porque en el proyecto
    original venía negra y se perdía sobre fondo oscuro.
    """
    if fig is None:
        return fig
    try:
        fig.update_layout(
            template=None,
            paper_bgcolor=DERIV_DARK_COLORS["paper"],
            plot_bgcolor=DERIV_DARK_COLORS["plot"],
            font=dict(color=DERIV_DARK_COLORS["text"], family="Arial"),
            legend=dict(
                bgcolor="rgba(0,0,0,0)",
                font=dict(color=DERIV_DARK_COLORS["text"]),
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1,
            ),
            margin=dict(l=70, r=50, t=72, b=80),
        )
        fig.update_xaxes(
            showgrid=False,
            zeroline=False,
            color=DERIV_DARK_COLORS["text"],
            title_font=dict(color=DERIV_DARK_COLORS["text"]),
            tickfont=dict(color=DERIV_DARK_COLORS["text"]),
            linecolor="rgba(255,255,255,0.20)",
        )
        fig.update_yaxes(
            showgrid=True,
            gridcolor=DERIV_DARK_COLORS["grid"],
            zeroline=False,
            color=DERIV_DARK_COLORS["text"],
            title_font=dict(color=DERIV_DARK_COLORS["text"]),
            tickfont=dict(color=DERIV_DARK_COLORS["text"]),
            linecolor="rgba(255,255,255,0.20)",
        )

        for trace in fig.data:
            name = str(getattr(trace, "name", "") or "").lower()
            if "total" in name:
                if hasattr(trace, "line"):
                    trace.line.color = DERIV_DARK_COLORS["cyan"]
                    trace.line.width = 3.2
                if hasattr(trace, "marker"):
                    try:
                        trace.marker.color = DERIV_DARK_COLORS["cyan"]
                    except Exception:
                        pass
            elif hasattr(trace, "line") and getattr(trace.line, "color", None) in {"black", "#000", "#000000", "rgb(0,0,0)"}:
                trace.line.color = DERIV_DARK_COLORS["primary"]
                trace.line.width = 2.4
    except Exception:
        pass
    return fig


def render_deriv_plotly(fig, key: str | None = None) -> None:
    """Renderiza Plotly con tema oscuro sin recursión accidental."""
    themed = apply_derivatives_dark_theme(fig)
    if key:
        st.plotly_chart(themed, width="stretch", key=key)
    else:
        st.plotly_chart(themed, width="stretch")



def module_group_name(block_name: str) -> str:
    """Etiqueta principal con prefijo D0x, como la navegación aprobada."""
    if block_name.startswith("D01"):
        return "D01 · Comparación"
    if block_name.startswith("D02"):
        return "D02 · FX USD/CLP"
    if block_name.startswith("D04"):
        return "D04 · SPC nominal"
    if block_name.startswith("D05"):
        return "D05 · UF/CLP"
    return "Otros"


def block_letter_label(block_name: str) -> str:
    """Etiqueta de submenú con letra cuando aplica."""
    return short_block_label(block_name)


def spec_label(spec) -> str:
    title = getattr(spec, "title", "") or getattr(spec, "module_name", "Gráfico")
    return re.sub(r"\s+", " ", str(title)).strip()

def block_group_name(block_name: str) -> str:
    """Compatibilidad con código heredado."""
    return module_group_name(block_name)


def short_block_label(block_name: str) -> str:
    cleaned = re.sub(r"^D\d{2}[A-Z]?(?:\s*-\s*)?", "", block_name).strip()
    return cleaned or block_name


def _append_export_frame(all_frames: list[pd.DataFrame], df_in: pd.DataFrame, spec_in) -> None:
    if df_in is None or df_in.empty:
        return
    tmp = df_in.copy()
    tmp.insert(0, "chart_module", spec_in.module_name)
    tmp.insert(1, "chart_title", spec_in.title)
    tmp.insert(2, "block", spec_in.block)
    all_frames.append(tmp)



def _password_hash(password: str) -> str:
    """Hash para invalidar cache si cambia la clave, sin guardar/imprimir la clave."""
    return hashlib.sha256((password or "").encode("utf-8")).hexdigest()


def _deriv_cache_key(start: date, end: date, active_map_path: str, user: str) -> str:
    map_path = Path(active_map_path)
    map_version = str(map_path.stat().st_mtime_ns) if map_path.exists() else "0"
    return f"{APP_VERSION}|{start}|{end}|{map_version}|{user}"


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def cached_preload_all_derivative_frames(start_date: str, end_date: str, active_map_path: str, map_version: str, app_version: str, user: str, password_hash: str, _password: str) -> dict[str, pd.DataFrame]:
    """Cache persistente de todos los dataframes de derivados.

    Streamlit rerun ocurre siempre al cambiar botones/radios, pero esta función
    evita recalcular/consultar los dataframes mientras no cambie:
    rango, mapa, versión, usuario o hash de clave.
    """
    frames: dict[str, pd.DataFrame] = {}
    for spec in CHARTS:
        try:
            df = cached_build_dataframe(
                spec.module_name,
                start_date,
                end_date,
                active_map_path,
                map_version,
                app_version,
                user,
                password_hash,
                _password,
            )
            frames[spec.module_name] = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
        except Exception as exc:
            frames[spec.module_name] = pd.DataFrame({"error": [str(exc)], "chart_module": [spec.module_name]})
    return frames


def preload_all_derivative_frames(start: date, end: date, active_map_path: str, user: str, password: str) -> dict[str, pd.DataFrame]:
    """Dataframes precargados/cacheados de derivados.

    Nota: Streamlit siempre hace rerun visual cuando cambia un radio/tab. Lo que
    evitamos acá es repetir consultas/cálculos de datos.
    """
    map_path = Path(active_map_path)
    map_version = str(map_path.stat().st_mtime_ns) if map_path.exists() else "0"
    pass_hash = _password_hash(password)
    key = f"{APP_VERSION}|{start}|{end}|{map_version}|{user}|{pass_hash}"
    expected_modules = {spec.module_name for spec in CHARTS}

    current = st.session_state.get("deriv_preloaded_frames")
    current_key = st.session_state.get("deriv_preload_key")
    if (
        current_key == key
        and isinstance(current, dict)
        and expected_modules.issubset(set(current.keys()))
    ):
        return current

    frames = cached_preload_all_derivative_frames(
        str(start),
        str(end),
        active_map_path,
        map_version,
        APP_VERSION,
        user,
        pass_hash,
        password,
    )

    st.session_state["deriv_preload_key"] = key
    st.session_state["deriv_preloaded_frames"] = frames
    return frames




def _infer_date_col(df: pd.DataFrame) -> str | None:
    for col in ["date", "fecha", "period", "month", "mes"]:
        if col in df.columns:
            return col
    for col in df.columns:
        s = str(col).lower()
        if "date" in s or "fecha" in s:
            return col
    return None


def _numeric_candidates(df: pd.DataFrame) -> list[str]:
    nums = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    blacklist = {"year", "month", "month_num", "order", "rank"}
    return [c for c in nums if str(c).lower() not in blacklist]


def _infer_value_col(df: pd.DataFrame) -> str | None:
    preferred = ["value", "valor", "amount", "monto", "stock", "flujo", "notional", "price", "precio"]
    for col in preferred:
        if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
            return col
    nums = _numeric_candidates(df)
    return nums[0] if nums else None


def _filter_frame(df: pd.DataFrame, filters: dict[str, str] | None = None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    for col, value in (filters or {}).items():
        if col in out.columns:
            out = out[out[col].astype(str).str.casefold().eq(str(value).casefold())].copy()
    return out


def _series_from_frame(df: pd.DataFrame, filters: dict[str, str] | None = None) -> pd.DataFrame:
    """Convierte un dataframe de gráfico en serie agregada por fecha."""
    df = _filter_frame(df, filters)
    if df is None or df.empty:
        return pd.DataFrame(columns=["date", "value"])
    date_col = _infer_date_col(df)
    value_col = _infer_value_col(df)
    if not date_col or not value_col:
        return pd.DataFrame(columns=["date", "value"])
    tmp = df[[date_col, value_col]].copy()
    tmp["date"] = pd.to_datetime(tmp[date_col], errors="coerce")
    tmp["value"] = pd.to_numeric(tmp[value_col], errors="coerce")
    tmp = tmp.dropna(subset=["date", "value"])
    if tmp.empty:
        return pd.DataFrame(columns=["date", "value"])
    return tmp.groupby("date", as_index=False)["value"].sum().sort_values("date")


def _metric_from_module(preloaded: dict[str, pd.DataFrame], module_name: str, filters: dict[str, str] | None = None) -> dict:
    series = _series_from_frame(preloaded.get(module_name, pd.DataFrame()), filters)
    if series.empty:
        return {"module": module_name, "date": None, "value": None, "prev": None, "var_abs": None, "var_pct": None}
    latest = series.iloc[-1]
    prev = series.iloc[-2] if len(series) >= 2 else None
    value = float(latest["value"])
    prev_value = float(prev["value"]) if prev is not None else None
    var_abs = value - prev_value if prev_value is not None else None
    var_pct = (var_abs / prev_value * 100.0) if prev_value not in (None, 0) else None
    return {
        "module": module_name,
        "date": pd.to_datetime(latest["date"]),
        "value": value,
        "prev": prev_value,
        "var_abs": var_abs,
        "var_pct": var_pct,
    }


def _deriv_state(var_pct: float | None, high_is_alert: bool = True) -> tuple[str, str]:
    if var_pct is None or pd.isna(var_pct):
        return "gris", "Sin comparación mensual"
    v = float(var_pct)
    move = abs(v)
    if move >= 20:
        return "rojo", "Cambio mensual extremo"
    if move >= 10:
        return "amarillo", "Cambio mensual relevante"
    return "verde", "Movimiento mensual acotado"


def _fmt_deriv_value(x: float | None, unit: str = "") -> str:
    if x is None or pd.isna(x):
        return "Sin dato"
    ax = abs(float(x))
    if ax >= 1_000_000:
        txt = f"{x/1_000_000:,.2f} MM"
    elif ax >= 1_000:
        txt = f"{x:,.0f}"
    else:
        txt = f"{x:,.2f}"
    txt = txt.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{txt} {unit}".strip()


def _fmt_pct(x: float | None) -> str:
    if x is None or pd.isna(x):
        return "Sin dato previo"
    return f"{x:+,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_date(x) -> str:
    if x is None or pd.isna(x):
        return "Sin fecha"
    return pd.to_datetime(x).strftime("%Y-%m")


def _summary_card(title: str, metric: dict, unit: str, detail: str, note: str = "") -> tuple[str, dict]:
    estado, mensaje = _deriv_state(metric.get("var_pct"))
    color_map = {"verde": "#20E080", "amarillo": "#FFD84D", "rojo": "#FF6B8A", "gris": "#94A3B8"}
    color = color_map.get(estado, color_map["gris"])
    value_txt = _fmt_deriv_value(metric.get("value"), unit)
    date_txt = _fmt_date(metric.get("date"))
    var_txt = _fmt_pct(metric.get("var_pct"))
    html = (
        f'<div class="kpi-card" style="border-left: 5px solid {color};">'
        f'<div class="kpi-top"><span class="kpi-title">{title}</span>'
        f'<span class="kpi-status" style="color:{color};">● {estado}</span></div>'
        f'<div class="kpi-value">{value_txt}</div>'
        f'<div class="kpi-meta">{detail} · último dato {date_txt}</div>'
        f'<div class="kpi-delta">Variación mensual: {var_txt}</div>'
        f'<div class="kpi-message">{mensaje if not note else note}</div>'
        f'</div>'
    )
    row = {
        "indicador": title,
        "estado": estado,
        "ultimo_dato": date_txt,
        "valor": metric.get("value"),
        "variacion_pct": metric.get("var_pct"),
        "mensaje": mensaje if not note else note,
        "modulo": metric.get("module"),
    }
    return html, row


def _render_report_section(title: str, intro: str, cards: list[tuple[str, dict]]) -> list[dict]:
    st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="muted-note">{intro}</div>', unsafe_allow_html=True)
    st.markdown('<div class="kpi-grid">' + "".join(card for card, _ in cards) + '</div>', unsafe_allow_html=True)
    return [row for _, row in cards]


def render_derivatives_executive_summary(preloaded: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Resumen ejecutivo alineado con el Informe Mensual BCCh/SIID.

    Estructura:
    I. Mercado total.
    II. Derivados y spot USD/CLP.
    III. SPC nominal.
    IV. UF-CLP.
    """
    all_rows: list[dict] = []

    # I. Mercado total: vigentes y transados por activo subyacente.
    sec1_cards = [
        _summary_card("Vigente total derivados", _metric_from_module(preloaded, "d01_resumen_stock", {"market": "Total"}), "MM USD", "I. Mercados de derivados"),
        _summary_card("Transado mensual total", _metric_from_module(preloaded, "d01_resumen_flujo", {"market": "Total"}), "MM USD", "I. Mercados de derivados"),
        _summary_card("Vigente tasas de interés", _metric_from_module(preloaded, "d01_resumen_stock", {"market": "Tasas de interés"}), "MM USD", "Subyacente de mayor tamaño"),
        _summary_card("Vigente tipos de cambio", _metric_from_module(preloaded, "d01_resumen_stock", {"market": "Tipos de cambio"}), "MM USD", "Subyacente FX"),
        _summary_card("Vigente UF/CLP", _metric_from_module(preloaded, "d01_resumen_stock", {"market": "UF/CLP"}), "MM USD", "Subyacente inflación"),
        _summary_card("Transado tasas", _metric_from_module(preloaded, "d01_resumen_flujo", {"market": "Tasas de interés"}), "MM USD", "Actividad mensual tasas"),
        _summary_card("Transado FX", _metric_from_module(preloaded, "d01_resumen_flujo", {"market": "Tipos de cambio"}), "MM USD", "Actividad mensual FX"),
        _summary_card("Transado UF/CLP", _metric_from_module(preloaded, "d01_resumen_flujo", {"market": "UF/CLP"}), "MM USD", "Actividad mensual UF/CLP"),
    ]
    all_rows += _render_report_section(
        "I. Mercados de derivados",
        "Sigue la lógica del informe: montos vigentes y montos transados por activo subyacente, expresados en dólares equivalentes.",
        sec1_cards,
    )

    # II. USD/CLP: derivados + spot, vigentes, transados, NDF.
    sec2_cards = [
        _summary_card("USD/CLP spot + derivados", _metric_from_module(preloaded, "d02_der_spot_total", {"dimension_1": "Total"}), "MM USD", "Mercado cambiario total"),
        _summary_card("Spot USD/CLP", _metric_from_module(preloaded, "d02_der_spot_total", {"dimension_1": "Spot"}), "MM USD", "Mercado spot"),
        _summary_card("Derivados USD/CLP transados", _metric_from_module(preloaded, "d02_der_spot_total", {"dimension_1": "Derivados"}), "MM USD", "Derivados cambiarios"),
        _summary_card("Derivados USD/CLP vigentes", _metric_from_module(preloaded, "d02_fx_vigente_instrumento"), "MM USD", "Montos vigentes por instrumento"),
        _summary_card("NDF no residentes vigente", _metric_from_module(preloaded, "d02_ndf_no_residentes_plazo"), "MM USD", "NDF por plazo contractual"),
        _summary_card("Spot por sector", _metric_from_module(preloaded, "d02_spot_sector"), "MM USD", "Contrapartes spot"),
    ]
    all_rows += _render_report_section(
        "II. Derivados y spot sobre tipos de cambio",
        "Resume el mercado USD/CLP separando spot, derivados, montos vigentes y NDF con no residentes.",
        sec2_cards,
    )

    # III. SPC nominal.
    sec3_cards = [
        _summary_card("SPC nominal transado", _metric_from_module(preloaded, "d04_spc_nominal_sector"), "MM CLP", "Transado por mercado/sector"),
        _summary_card("SPC nominal vigente", _metric_from_module(preloaded, "d04_spc_nominal_vigente_sector"), "MM CLP", "Vigente por mercado/sector"),
        _summary_card("SPC transado por plazo", _metric_from_module(preloaded, "d04_spc_nominal_plazo"), "MM CLP", "Plazo contractual hasta 2 años"),
        _summary_card("SPC vigente por plazo", _metric_from_module(preloaded, "d04_spc_nominal_vigente_plazo"), "MM CLP", "Plazo residual hasta 2 años"),
    ]
    all_rows += _render_report_section(
        "III. Derivados SPC nominal",
        "Replica el foco del informe en SPC-CLP hasta dos años, útil para leer actividad y posiciones de tasas locales.",
        sec3_cards,
    )

    # IV. UF-CLP.
    sec4_cards = [
        _summary_card("Forward UF/CLP transado", _metric_from_module(preloaded, "d05_ufclp_forward_sector"), "MM UF", "Forward mercado local"),
        _summary_card("Swap UF/CLP transado", _metric_from_module(preloaded, "d05_ufclp_swap_sector"), "MM UF", "Swap UF/CLP"),
        _summary_card("Precio forward UF/CLP 12M", _metric_from_module(preloaded, "d05_ufclp_forward_prices"), "CLP/UF", "Precio promedio 12 meses"),
    ]
    all_rows += _render_report_section(
        "IV. Derivados sobre UF-CLP",
        "Sigue la sección de inflación/UF del informe: forwards, swaps y precios forward UF/CLP.",
        sec4_cards,
    )

    summary = pd.DataFrame(all_rows)
    with st.expander("Ver tabla resumen derivados", expanded=False):
        st.dataframe(summary, width="stretch", hide_index=True)
    st.markdown(
        '<div class="muted-note">Lectura: verde = variación acotada; amarillo = cambio mensual relevante; rojo = cambio mensual extremo; gris = sin comparación. '
        'Los umbrales son ejecutivos y pueden ajustarse luego por criterio regulatorio.</div>',
        unsafe_allow_html=True,
    )
    return summary


def render_derivatives_module(start: date, end: date, user: str, password: str) -> dict[str, pd.DataFrame]:
    """Renderiza Derivados con dos páginas internas:
    - 8.0 Resumen
    - 8.1 Derivados
    """
    set_derivatives_credentials(user, password)

    section_title("Derivados bancarios")

    if _DERIV_IMPORT_ERROR is not None:
        st.error(
            "El módulo de derivados no está instalado completo. "
            "Debes subir también la carpeta derivatives_src/ y derivatives_app/config/. "
            f"Detalle interno: {_DERIV_IMPORT_ERROR}"
        )
        return {"derivados_error": pd.DataFrame({"error": [str(_DERIV_IMPORT_ERROR)]})}

    st.caption(
        "Proyecto mensual de derivados BCCh/SIID integrado al monitor. "
        "Usa las mismas credenciales SieteWS activas del monitor principal y mantiene las rutas BDE oficiales del módulo original."
    )

    if not DEFAULT_MAP.exists():
        st.error(f"No existe el archivo de mapeo de derivados: {DEFAULT_MAP}")
        return {"derivados_error": pd.DataFrame({"error": [f"No existe {DEFAULT_MAP}"]})}

    active_map_path = str(DEFAULT_MAP)
    registry = load_registry(active_map_path)
    block_list = blocks()

    if not block_list:
        st.warning("No hay bloques de derivados disponibles.")
        return {"derivados_vista_actual": pd.DataFrame()}

    # Datos cacheados/precargados para ambas páginas.
    preloaded = preload_all_derivative_frames(start, end, active_map_path, user, password)

    # Página interna: radio horizontal para evitar que Streamlit renderice ambas páginas a la vez.
    st.markdown('<div class="muted-note"><b>Vista derivados</b></div>', unsafe_allow_html=True)
    deriv_page = st.radio(
        "Vista derivados",
        options=["8.0 Resumen", "8.1 Derivados"],
        index=0,
        horizontal=True,
        label_visibility="collapsed",
        key="deriv_internal_page",
    )

    if deriv_page == "8.0 Resumen":
        summary_df = render_derivatives_executive_summary(preloaded)
        st.markdown(
            '<div class="muted-note">El resumen es una capa ejecutiva por bloque. '
            'Para ver todos los gráficos originales, entra a <b>8.1 Derivados</b>.</div>',
            unsafe_allow_html=True,
        )
        return {"derivados_resumen": summary_df}

    # 8.1 Derivados: navegación original aprobada.
    all_frames: list[pd.DataFrame] = []

    group_to_blocks: dict[str, list[str]] = {}
    for _block in block_list:
        group_to_blocks.setdefault(module_group_name(_block), []).append(_block)

    group_names = list(group_to_blocks.keys())

    st.markdown('<div class="muted-note"><b>1. Módulo</b></div>', unsafe_allow_html=True)
    selected_group = st.radio(
        "Módulo derivados",
        options=group_names,
        index=0,
        horizontal=True,
        label_visibility="collapsed",
        key="deriv_module_group_radio",
    )

    group_blocks = group_to_blocks[selected_group]
    st.markdown('<div class="muted-note" style="margin-top:.65rem;"><b>2. Submenú</b></div>', unsafe_allow_html=True)
    selected_block = st.radio(
        "Submenú derivados",
        options=group_blocks,
        index=0,
        horizontal=True,
        format_func=block_letter_label,
        label_visibility="collapsed",
        key="deriv_module_submenu_radio",
    )

    b1, b2, b3 = st.columns([1.0, 1.0, 4.0], gap="large")
    with b1:
        if st.button("Probar BDE", width="stretch", key="deriv_test_bde"):
            try:
                rows, periods = bde_test_count()
                st.success(f"BDE OK: {rows:,} filas y {periods:,} períodos.")
            except Exception as exc:
                st.error(f"Falló BDE derivados: {exc}")
    with b2:
        if st.button("Limpiar cache", width="stretch", key="deriv_clear_cache"):
            st.session_state.pop("deriv_preload_key", None)
            st.session_state.pop("deriv_preloaded_frames", None)
            st.cache_data.clear()
            st.rerun()

    specs = charts_for_block(selected_block)

    st.markdown(
        f"""
        <div class="reading-box">
            <b>Vista actual:</b> {selected_block}<br>
            <span style="color: var(--cmf-muted);">Se muestran todos los gráficos de esta vista: {len(specs)} · Dataframes precargados: {len(preloaded)}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    def append_frame(df_in: pd.DataFrame, spec_in) -> None:
        if df_in is None or df_in.empty:
            return
        tmp = df_in.copy()
        tmp.insert(0, "chart_module", getattr(spec_in, "module_name", ""))
        tmp.insert(1, "chart_title", getattr(spec_in, "title", ""))
        tmp.insert(2, "block", getattr(spec_in, "block", selected_block))
        all_frames.append(tmp)

    # Se mantienen los renderizadores especiales del proyecto original.
    if selected_block.startswith("D01"):
        spec_map = {s.module_name: s for s in specs}
        for spec in specs:
            if spec.module_name in {"d01_waterfall_variacion", "d01_variacion_componentes"}:
                continue
            df_chart = render_chart(spec, start, end, active_map_path)
            append_frame(df_chart, spec)
        if "d01_waterfall_variacion" in spec_map and "d01_variacion_componentes" in spec_map:
            pair_frames = render_d01_variation_pair(
                spec_map["d01_waterfall_variacion"],
                spec_map["d01_variacion_componentes"],
                start,
                end,
                active_map_path,
            )
            for spec_name, df_pair in zip(["d01_waterfall_variacion", "d01_variacion_componentes"], pair_frames):
                append_frame(df_pair, spec_map[spec_name])

    elif selected_block.startswith("D02"):
        spec_map = {s.module_name: s for s in specs}

        if "d02_ndf_no_residentes_plazo" in spec_map and "d02_ndf_variacion_plazo" in spec_map:
            ndf_frames = render_d02_ndf_pair(
                spec_map["d02_ndf_no_residentes_plazo"],
                spec_map["d02_ndf_variacion_plazo"],
                start,
                end,
                active_map_path,
                "d02_ndf_vig",
                "NDF vigentes netos por plazo contractual",
            )
            for df_pair, spec_name in zip(ndf_frames, ["d02_ndf_no_residentes_plazo", "d02_ndf_variacion_plazo"]):
                append_frame(df_pair, spec_map[spec_name])

        elif "d02_ndf_transados_plazo" in spec_map and "d02_ndf_transados_variacion_plazo" in spec_map:
            ndf_frames = render_d02_ndf_pair(
                spec_map["d02_ndf_transados_plazo"],
                spec_map["d02_ndf_transados_variacion_plazo"],
                start,
                end,
                active_map_path,
                "d02_ndf_tra",
                "NDF transados netos por plazo contractual",
            )
            for df_pair, spec_name in zip(ndf_frames, ["d02_ndf_transados_plazo", "d02_ndf_transados_variacion_plazo"]):
                append_frame(df_pair, spec_map[spec_name])

        else:
            for spec in specs:
                if spec.module_name in {"d02_fx_variacion_apertura_sector", "d02_spot_compras_ventas_variacion"}:
                    continue

                df_chart = render_chart(spec, start, end, active_map_path)
                append_frame(df_chart, spec)

                if spec.module_name == "d02_fx_vigente_contraparte" and "d02_fx_variacion_apertura_sector" in spec_map:
                    pair_frames = render_d02_variacion_apertura(
                        spec_map["d02_fx_variacion_apertura_sector"], start, end, active_map_path
                    )
                    for df_pair in pair_frames:
                        append_frame(df_pair, spec_map["d02_fx_variacion_apertura_sector"])

                if spec.module_name == "d02_spot_sector" and "d02_spot_compras_ventas_variacion" in spec_map:
                    spot_cv_frames = render_d02_spot_compras_ventas_variacion(
                        spec_map["d02_spot_compras_ventas_variacion"], start, end, active_map_path
                    )
                    for df_pair in spot_cv_frames:
                        append_frame(df_pair, spec_map["d02_spot_compras_ventas_variacion"])

    elif selected_block.startswith("D04-A"):
        d04_frames = render_d04_spc_nominal(specs, start, end, active_map_path)
        for df_chart, spec in zip(d04_frames, specs):
            append_frame(df_chart, spec)

    elif selected_block.startswith("D04-B"):
        d04_frames = render_d04_spc_nominal_vigentes(specs, start, end, active_map_path)
        for df_chart, spec in zip(d04_frames, specs):
            append_frame(df_chart, spec)

    elif selected_block.startswith("D05-A"):
        d05_frames = render_d05_ufclp_transados(specs, start, end, active_map_path)
        for df_chart, spec in zip(d05_frames, specs):
            append_frame(df_chart, spec)

    else:
        for i, spec in enumerate(specs):
            df_chart = render_chart(spec, start, end, active_map_path, add_divider=i < len(specs) - 1)
            append_frame(df_chart, spec)

    export_df = pd.concat(all_frames, ignore_index=True, sort=False) if all_frames else pd.DataFrame()

    with st.expander("Exportación derivados", expanded=False):
        if not export_df.empty:
            st.download_button(
                "Descargar datos de la vista actual CSV",
                data=export_df.to_csv(index=False, encoding="utf-8-sig"),
                file_name="derivados_vista_actual.csv",
                mime="text/csv",
                width="stretch",
                key="deriv_export_visible_csv",
            )

        if st.button("Preparar Excel Power BI de todos los gráficos de derivados", width="stretch", key="deriv_prepare_powerbi"):
            try:
                excel_bytes = build_powerbi_excel_for_download(start, end, active_map_path)
                st.session_state["deriv_powerbi_excel"] = excel_bytes
                st.success("Excel preparado.")
            except Exception as exc:
                st.error(f"No se pudo preparar Excel Power BI: {exc}")

        if "deriv_powerbi_excel" in st.session_state:
            st.download_button(
                "Descargar Excel Power BI derivados",
                data=st.session_state["deriv_powerbi_excel"],
                file_name="bcch_derivados_mensual_graficos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
                key="deriv_download_powerbi",
            )

    st.caption("Módulo derivados integrado. Siempre muestra todos los gráficos del submenú seleccionado.")
    return {"derivados_vista_actual": export_df}
